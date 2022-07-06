import openpyxl
from pathlib import Path


xlsx_files = [path for path in Path('.').rglob('*.xlsx')]
for a in xlsx_files:
    print(a)

print("---------")
#xlsx_file = Path('.', 'Cliente1-Obra1-15-jul-22.xlsx')
#wb_obj = openpyxl.load_workbook(xlsx_file, data_only=True) 


wbs = [openpyxl.load_workbook(wb, data_only=True) for wb in xlsx_files]
for wb in wbs:    
    # Read the active sheet:
    sheet = wb.active
    currTotal = 0
    for row in sheet.iter_rows(max_row=30):
        for cell in row:
            if cell.value is not None :
                if cell.column == 9:
                    currTotal = cell.value

    print(currTotal)
